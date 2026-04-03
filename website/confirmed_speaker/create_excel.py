import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Confirmed Speakers"

# Define styles
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='Calibri', size=11)
data_alignment = Alignment(vertical='top', wrap_text=True)
bio_alignment = Alignment(vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

alt_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

# Headers
headers = ['S.No.', 'Name', 'Designation', 'Organization', 'Email', 'Brief Bio']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Speaker data with bios
speakers = [
    {
        'name': 'Victoria Goliber',
        'designation': 'Director of Research Partnerships and Government Programs',
        'org': 'D-Wave Systems',
        'email': 'vgoliber@dwavesys.com',
        'bio': (
            "Dr. Victoria Goliber leads government grants strategy and research partnerships at D-Wave Quantum, the world's first commercial supplier of quantum computers. "
            "She holds a Ph.D. in Discrete Mathematics from Arizona State University (U.S. DoD SMART Scholar) and an M.S. in Computer Science (Machine Learning) from Georgia Tech. "
            "Previously a Senior Mathematician at the U.S. Air Force Research Laboratory, she brings a unique defense-to-industry perspective. "
            "Victoria is passionate about demystifying quantum computing and building quantum-ready workforce pipelines globally."
        )
    },
    {
        'name': 'Ritajit Majumdar',
        'designation': 'Research Scientist',
        'org': 'IBM Quantum (IBM India Research Lab)',
        'email': 'majumdar.ritajit@gmail.com',
        'bio': (
            "Dr. Ritajit Majumdar is a Research Scientist at IBM Quantum, enabling utility-scale quantum results on IBM devices. "
            "He earned his Ph.D. from the Indian Statistical Institute specializing in quantum error correction, supported by a Fulbright-Nehru Doctoral Research Fellowship at IBM Watson Research Centre. "
            "A gold medalist from Calcutta University and DST Inspire Fellow, he has published extensively on quantum error mitigation, NISQ algorithms, and machine learning-based decoding. "
            "Ritajit is a sought-after speaker who enjoys bridging the gap between quantum theory and practical applications."
        )
    },
    {
        'name': 'Randy Kuang',
        'designation': 'Co-founder & Chief Scientist',
        'org': 'Quantropi Inc.',
        'email': 'randy.kuang@quantropi.com',
        'bio': (
            "Dr. Randy Kuang is the Co-founder and Chief Scientist of Quantropi, building quantum-safe cryptographic solutions for the post-quantum era. "
            "He holds a doctorate in Quantum Physics, and NASA named his work \"Kuang's semi-classical formalism\" in 2012. "
            "A prolific inventor with 30+ U.S. patents spanning WiMAX, optical networks, and quantum cryptography, he previously served as Senior Network Researcher at Nortel and CTO at inBay Technologies. "
            "Randy's pioneering Quantum Permutation Pad (QPP) encryption has been demonstrated on IBM quantum systems — he loves turning deep theory into deployable security."
        )
    },
    {
        'name': 'Dr. Shravani Shahapure',
        'designation': 'Associate Director – Quantum Safe Cybersecurity',
        'org': 'Big4 (Risk Advisory)',
        'email': 'shra.shahapure@gmail.com',
        'bio': (
            "Dr. Shravani Shahapure is an Associate Director at Big4, leading the Quantum Safe cybersecurity, Cryptography, and Quantum Computing practice within Risk Advisory. "
            "With 20+ years in cybersecurity, she holds two patents in cryptography and quantum computing and an M.E. from IIT Bombay."
            "She previously held senior roles at Capgemini, PwC India, and the Data Security Council of India. "
            "Shravani is deeply committed to helping enterprises navigate the quantum threat landscape and is known for making complex security concepts accessible."
        )
    },
    {
        'name': 'Biman Chattopadhyay',
        'designation': 'Co-founder & CTO',
        'org': 'Quanfluence',
        'email': 'biman@quanfluence.com',
        'bio': (
            "Biman Chattopadhyay is the Co-founder and CTO of Quanfluence, a Bengaluru-based quantum computing startup building India's first general-purpose quantum computer using photonic technology. "
            "He leads the development of their optical Ising machine that processes hundreds of interconnected variables simultaneously for optimization problems. "
            "Quanfluence recently secured $2M in seed funding, and Biman holds patents in quantum random number generation. "
            "An entrepreneur at heart, he also co-founded SilabTech and is an active angel investor in deep-tech startups."
        )
    },
    {
        'name': 'Dr. Sourav Chatterjee',
        'designation': 'Research Engineer',
        'org': 'TCS Innovation Labs (TCS Research)',
        'email': 'sourav.chat@tcs.com',
        'bio': (
            "Dr. Sourav Chatterjee is a Research Engineer at TCS Innovation Labs, Pune, specializing in quantum computing, quantum sensing, and quantum communication. "
            "He is a doctoral candidate in Physics with expertise in photonic Ising machines, secure quantum communication, and quantum metrology. "
            "His work at TCS explores quantum sensing for industrial applications in healthcare, electronics, and magnetic material analysis. "
            "Sourav is driven by curiosity and enjoys finding practical industrial use cases for quantum technologies that can create real-world impact."
        )
    },
    {
        'name': 'Dr. Nayana Das',
        'designation': 'Research Engineer',
        'org': 'LTI Mindtree',
        'email': 'dasnayana92@gmail.com',
        'bio': (
            "Dr. Nayana Das is a Research Engineer at LTI Mindtree specializing in quantum communication and post-quantum cryptography, with 9+ years of experience in quantum cryptography and information security. "
            "She earned her Ph.D. in Computer Science from the Indian Statistical Institute with a thesis on \"Analysis and Design of Quantum Secure Communication Systems.\" "
            "She holds multiple patents and has published in prestigious journals on topics including Measurement-Device-Independent Quantum Secure Direct Communication. "
            "Nayana is passionate about building secure communication infrastructure for the quantum era and mentoring the next generation of researchers."
        )
    },
    # {
    #     'name': 'Dr. Rashi Koul',
    #     'designation': 'Quantum Technology Expert',
    #     'org': 'Taqbit Labs',
    #     'email': 'rashi@taqbit.com',
    #     'bio': (
    #         "Dr. Rashi Koul is a quantum technology expert at Taqbit Labs, a company specializing in quantum-safe cybersecurity solutions including Quantum Key Distribution (QKD) and Quantum Random Number Generators (QRNG). "
    #         "She contributes to Taqbit's mission of building quantum-secure communication systems for defense, telecom, and enterprise infrastructure. "
    #         "With a doctoral background, she brings strong research capabilities to the development of practical quantum security solutions. "
    #         "Rashi is motivated by the challenge of making quantum-safe technologies accessible and deployable for organizations preparing for the post-quantum world."
    #     )
    # },
    {
        'name': 'Mr. Animesh Aaryan',
        'designation': 'Founder & CEO',
        'org': 'Taqbit Labs',
        'email': 'animesh@taqbit.com',
        'bio': (
            "Mr. Animesh Aaryan stands at the forefront of the quantum technology revolution as the Chief Executive Officer of TAQBIT LABS, Bengaluru. "
            "He is an innovator with close to a decade of experience in quantum physics, with a strong focus on quantum cryptography and photonics. "
            "TAQBIT Labs is a technology company focused on commercializing Quantum technologies such as Quantum Communication, Sensing, & Algorithms, which works very closely with the Ministry of Defence, Government of India. "
            "In recognition of its impact and innovation, TAQBIT Labs was awarded Startup of the Year 2025."
        )
    },
    {
        'name': 'Dr. Jagrati Dwivedi',
        'designation': 'Scientist / Program Associate',
        'org': 'National Quantum Mission (NQM)',
        'email': 'jdwivedi.phy@gmail.com',
        'bio': (
            "Dr. Jagrati Dwivedi is associated with India's National Quantum Mission (NQM), the ₹6,003 crore government initiative to position India as a global leader in quantum science and technology. "
            "She holds a Ph.D. in Physics with expertise in materials science, X-ray diffraction, thin film deposition, and magnetic materials. "
            "Her research background spans advanced experimental physics with postdoctoral experience at DESY, Hamburg. "
            "Jagrati brings a unique blend of experimental physics expertise and policy perspective to India's quantum technology ecosystem."
        )
    },
    {
        'name': 'Dr. Pranab Dutta',
        'designation': 'CEO & Co-founder',
        'org': 'GDQLabs',
        'email': 'gdqlabs@gdqlabs.com',
        'bio': (
            "Dr. Pranab Dutta is the CEO and Co-founder of GDQLabs, specializing in quantum sensing for biological and medical applications. "
            "He holds a Ph.D. in Atomic Physics with 7+ years of experience in quantum technologies, and co-founded GDQLabs with fellow PhD physicists from IISER Pune. "
            "He led the development of India's first quantum gravimeter and the MAGSENSE™ quantum magnetic sensor for radiation-free cardiac diagnostics in under 5 minutes. "
            "Pranab is driven by the vision of bringing quantum precision to healthcare and making advanced diagnostics accessible to everyone."
        )
    },
    {
        'name': 'Dr. Amit Kumar Chauhan',
        'designation': 'Senior Research Associate',
        'org': 'QNu Labs Pvt. Ltd.',
        'email': 'amit.c@qnulabs.com',
        'bio': (
            "Dr. Amit Kumar Chauhan is a Senior Research Associate at QNu Labs, India's pioneering quantum cryptography company focused on quantum-safe security solutions. "
            "He holds a Ph.D. in Cryptography from IIT Ropar and an M.Tech in Computer Science from IIIT Delhi, with research interests in Post-Quantum Cryptography, Cryptanalysis, and Quantum Computation. "
            "Prior to QNu Labs, he worked as a Research Engineer in the PQC group at C-DOT (Centre for Development of Telematics), contributing to India's telecom security infrastructure. "
            "Amit is passionate about building cryptographic defenses that will stand the test of quantum computing and enjoys mentoring young researchers in the field."
        )
    },
    {
        'name': 'Dr. Anjani Priyadarsini',
        'designation': 'Quantum Industry Specialist',
        'org': 'Executive Advisor (Independent)',
        'email': 'anjani@example.com',
        'bio': (
            "Dr. Anjani Priyadarsini is a Quantum Industry Specialist and Executive Advisor. "
            "With extensive experience in Quantum Business Development (previously with AWS India), "
            "she is passionate about bridging the gap between quantum research and practical industry applications. She "
            "actively engages with academic institutions, enterprises, and the broader tech "
            "community to accelerate the adoption and understanding of quantum technologies "
            "in India and globally."
        )
    },
    {
        'name': 'Reena Dayal',
        'designation': 'Founder and CEO',
        'org': 'QETCI',
        'email': 'reena@qetci.org',
        'bio': (
            "Prominent technology leader, innovation expert, CEO and Partner for Benzaiten Advisors. "
            "Involved in IEEE Quantum Initiative, Government of India's Quantum Task Force, and Government of Telangana's Quantum Committee. "
            "Member of World Economic Forum's Global Futures Council for Quantum. Formerly founding Director for Microsoft Garage in India. "
            "Holds an Engineering degree from IIT Roorkee and is a Chevening Fellow."
        )
    },
    {
        'name': 'Dr. Kazuya Niizeki',
        'designation': 'CEO & Co-founder',
        'org': 'LQUOM',
        'email': 'kazuya.niizeki@lquom.com',
        'bio': (
            "CEO and founder of LQUOM Inc., a quantum technology startup focused on quantum communication systems, specifically quantum repeaters for a secure quantum internet. "
            "Recognized in Forbes' 2023 30 Under 30 - Asia list. Background from Yokohama National University."
        )
    },
]

# Write data
for idx, speaker in enumerate(speakers, 1):
    row = idx + 1
    fill = alt_fill if idx % 2 == 0 else None

    cells = [
        (1, idx),
        (2, speaker['name']),
        (3, speaker['designation']),
        (4, speaker['org']),
        (5, speaker['email']),
        (6, speaker['bio']),
    ]

    for col, value in cells:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = data_font
        cell.alignment = bio_alignment if col == 6 else data_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill

# Set column widths
col_widths = [6, 25, 40, 30, 30, 80]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Set row heights for data rows
for row in range(2, len(speakers) + 2):
    ws.row_dimensions[row].height = 80

# Freeze the header row
ws.freeze_panes = 'A2'

# Save
output_path = '/Users/quantsha/Downloads/what_QI_want/confirmed_speaker/confirmed_speakers.xlsx'
wb.save(output_path)
print(f"Excel file created at: {output_path}")
